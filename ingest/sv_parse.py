#!/usr/bin/env python3
"""Raw SwingVision parser (run with /usr/bin/python3 — has openpyxl).
Reads a match .xlsx 'Shots' sheet into per-shot dicts with real court
coordinates (meters), contact height, speed, spin, direction and timing —
the full data the zone-based Firestore export threw away."""
import openpyxl, sys, os, json

# court dimensions (meters): length 23.77, singles width 8.23 (x in [-4.115, 4.115]), net at y=11.885
COURT_L, COURT_W = 23.77, 8.23

def read_shots(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if 'Shots' not in wb.sheetnames:
        return None, None
    ws = wb['Shots']
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else '' for c in next(it)]
    idx = {h: i for i, h in enumerate(hdr)}
    def g(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None
    def fnum(row, col):
        v = g(row, col)
        try: return float(v)
        except (TypeError, ValueError): return None
    shots = []
    for row in it:
        if g(row, 'Player') in (None, '', 'None'): continue
        shots.append(dict(
            player=g(row, 'Player'), shotType=g(row, 'Type'), stroke=g(row, 'Stroke'),
            spin=g(row, 'Spin'), speed=fnum(row, 'Speed (MPH)'),
            point=g(row, 'Point'), game=g(row, 'Game'), set=g(row, 'Set'),
            bounceDepth=g(row, 'Bounce Depth'), bounceZone=g(row, 'Bounce Zone'), bounceSide=g(row, 'Bounce Side'),
            bx=fnum(row, 'Bounce (x)'), by=fnum(row, 'Bounce (y)'),
            hitDepth=g(row, 'Hit Depth'), hitZone=g(row, 'Hit Zone'), hitSide=g(row, 'Hit Side'),
            hx=fnum(row, 'Hit (x)'), hy=fnum(row, 'Hit (y)'), hz=fnum(row, 'Hit (z)'),
            direction=g(row, 'Direction'), result=g(row, 'Result'),
            startTime=g(row, 'Start Time'), videoTime=fnum(row, 'Video Time'),
        ))
    # settings for player names
    meta = {}
    if 'Settings' in wb.sheetnames:
        sit = wb['Settings'].iter_rows(values_only=True)
        sh = [str(c).strip() if c else '' for c in next(sit)]
        sr = next(sit, None)
        if sr:
            si = {h: i for i, h in enumerate(sh)}
            meta = {'host': sr[si.get('Host Team', 0)], 'guest': sr[si.get('Guest Team', 1)],
                    'start': sr[si.get('Start Time', 0)]}
    return shots, meta

if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else 'raw/rudy_youcef_sv.xlsx'
    shots, meta = read_shots(path)
    print('meta:', meta)
    print('shots:', len(shots))
    # quick rich-metric sanity per player
    from collections import defaultdict
    agg = defaultdict(lambda: {'n': 0, 'serveH': [], 'gsH': [], 'speed': [], 'contactY': []})
    for s in shots:
        a = agg[s['player']]; a['n'] += 1
        if s['stroke'] == 'Serve' and s['hz']: a['serveH'].append(s['hz'])
        if s['stroke'] in ('Forehand', 'Backhand') and s['hz']: a['gsH'].append(s['hz'])
        if s['speed']: a['speed'].append(s['speed'])
        if s['hy'] is not None: a['contactY'].append(s['hy'])
    for p, a in agg.items():
        avg = lambda L: round(sum(L) / len(L), 2) if L else None
        print(f"  {p:22} shots={a['n']:>3}  serve-contact-ht={avg(a['serveH'])}m  gs-contact-ht={avg(a['gsH'])}m  "
              f"avg-speed={avg(a['speed'])}mph  avg-contact-y={avg(a['contactY'])}m")
    print('\nsample groundstroke:', json.dumps(next((s for s in shots if s['stroke'] == 'Forehand'), {}), default=str)[:300])
